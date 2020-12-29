import os
import re
import pprint # pretty printing
import openpyxl # Excel Spreadsheet creation

# Change current working directory to file location
absPath = os.path.abspath(__file__)
dirName = os.path.dirname(absPath)
os.chdir(dirName + '\\Financial_Scraper\\text_file_data')

# Section 1: Pulling and Parsing Data
listings = []
listingsRegex = re.compile(r'(\d?\d\/\d\d\s.*)')

# Open and read file
with open('data.txt', 'r', encoding='UTF-8') as file:
    for line in file:
        # Debug: Print Lines
        #print(line)
        listing = listingsRegex.findall(line)

        # If the listing isn't blank, append it to listings 
        if listing:
            listings.append(listing)

# Separate listings into groups
strippedListings = []
for listing in listings:
    strippedListings.append(listing[0].split())

# Recombine listings into proper format
recombinedListings = []
for listing in strippedListings:
    # Dates
    date = listing[0]
    del listing[0]
    # Amounts
    amount = listing[-1]
    del listing[-1]
    # Descriptions
    description = ' '.join(listing)

    recombinedListings.append([date, description, amount])

# Separate Revenues from Expenses
revenues = []
expenses = []

for listing in recombinedListings:
    if listing[-1][-1] == '-':
        listing[-1] = listing[-1][:-1] # Strip '-' delimeter from end of expense value
        expenses.append(listing)
        print(listing)
    else:
        revenues.append(listing)

# Print values to spreadsheet
# Change directory to xlsx location
os.chdir(dirName + '\\Financial_Scraper\\final_xlsx_output')
# Open workbook
wb = openpyxl.load_workbook(os.getcwd() + '\\final_output.xlsx')

# Set Sheets
rSheet = wb['Revenues']
eSheet = wb['Expenses']


# Print revenue data to spreadsheet
for cols in range(0,3): # Outer loop for cols
    for rows in range(len(revenues)):
        rSheet.cell(row = rows + 2, column = cols + 1, value = revenues[rows][cols])

#Print expense data to spreadsheet
for cols in range(0,3): # Outer loop for cols
    for rows in range(len(expenses)):
        eSheet.cell(row = rows + 2, column = cols + 1, value = expenses[rows][cols])

# Save spreadsheet
wb.save('final_output.xlsx')